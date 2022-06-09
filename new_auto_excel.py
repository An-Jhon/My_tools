import openpyxl
import os
import sys
from openpyxl import load_workbook   # 读取表的模块

base_path = sys.path[0]
finaly_workbook_week = load_workbook(base_path + '\\表格\\周平衡计分卡.xlsx')
finaly_workbook_month = load_workbook(base_path + '\\表格\\月平衡计分卡.xlsx')


# take all tables
def take_tables(data_path):
    dirs = os.listdir(data_path)  # 获取路径下所有文件和文件夹名称
    is_exist_dir = []
    for i in dirs:
        if os.path.isdir(data_path+'\\'+i):
            is_exist_dir.append(i)

    if len(is_exist_dir) > 0:
        week_data_dirs_excels = []
        for dir in dirs:
            excel_names = os.listdir(data_path + dir)
            week_data_excels = []
            for name in excel_names:
                table = load_workbook(data_path + dir + "\\"+name)
                week_data_excels.append(table)
            week_data_dirs_excels.append(week_data_excels)

        return week_data_dirs_excels

    else:
        month_excels = []
        for month_table_name in dirs:
            month_table = load_workbook(data_path + "\\"+month_table_name)
            month_excels.append(month_table)

        return month_excels


# the review_data
def table_to_sheet(tables, finaly_table_sheet):
    for table in tables:
        # take the sheet
        table_sheet = table[table.sheetnames[0]]

        for i in range(1, finaly_table_sheet.max_row + 1):
            for j in range(1, table_sheet.max_row + 1):
                for k in range(1, table_sheet.max_column + 1):

                    # if the cell is empty, fill it with 0
                    if not table_sheet.cell(row=j, column=k).value:
                        table_sheet.cell(row=j, column=k, value=0)

                    # 判断总表的字段与审核概况表字段是否相同，相同则将分表数据填入总表中
                    if finaly_table_sheet.cell(row=i, column=1).value == table_sheet.cell(row=j, column=1).value and finaly_table_sheet.cell(row=i, column=2).value == table_sheet.cell(row=1, column=k).value:
                        finaly_table_sheet.cell(
                            row=i, column=3, value=table_sheet.cell(row=j, column=k).value)


# the AI_data
def table_to_AIsheet(tables, finaly_table_sheet):
    table0_sheet = tables[0][tables[0].sheetnames[0]]
    table1_sheet = tables[1][tables[1].sheetnames[0]]

    for i in range(1, finaly_table_sheet.max_row+1):
        for j in range(1, table1_sheet.max_row+1):
            for k in range(1, table1_sheet.max_column+1):
                if finaly_table_sheet.cell(row=i, column=2).value == table1_sheet.cell(row=j, column=2).value and finaly_table_sheet.cell(row=i, column=3).value == table1_sheet.cell(row=1, column=k).value:
                    finaly_table_sheet.cell(
                        row=i, column=5, value=table1_sheet.cell(row=j, column=k).value)

        for n in range(1, table0_sheet.max_column+1):
            if finaly_table_sheet.cell(row=i, column=3).value == table0_sheet.cell(row=1, column=n).value:
                finaly_table_sheet.cell(
                    row=i, column=5, value=table0_sheet.cell(row=2, column=n).value)

        if not finaly_table_sheet.cell(row=i, column=5).value:
            finaly_table_sheet.cell(
                row=i, column=5, value=finaly_table_sheet.cell(row=i, column=4).value)


# the week_data
def week_data():
    # 读取汇总表的三个sheet
    sheet1_finaly_workbook_week = finaly_workbook_week[finaly_workbook_week.sheetnames[0]]
    sheet2_finaly_workbook_week = finaly_workbook_week[finaly_workbook_week.sheetnames[1]]
    sheet3_finaly_workbook_week = finaly_workbook_week[finaly_workbook_week.sheetnames[2]]

    # take the table for filanySheet
    week_data_base_path = sys.path[0] + '\\周数据表\\'
    all_tables = take_tables(week_data_base_path)
    sheet1_excels = all_tables[0]
    sheet2_excels = all_tables[1]
    sheet3_excels = all_tables[2]

    # 开始匹配填表
    table_to_sheet(sheet1_excels, sheet1_finaly_workbook_week)
    print('sheet1填写保存完毕！！！')
    table_to_AIsheet(sheet2_excels, sheet2_finaly_workbook_week)
    print('sheet2填写保存完毕！！！')
    table_to_sheet(sheet3_excels, sheet3_finaly_workbook_week)
    print('sheet3填写保存完毕！！！')
    finaly_workbook_week.save(filename=sys.path[0] + '\\表格\\周平衡计分卡.xlsx')
    print('周数据全部填写完成！')


# the month_data
def month_data():
    sheet_finaly_workbook_month = finaly_workbook_month[finaly_workbook_month.sheetnames[0]]
    month_data_base_path = sys.path[0] + '\\月数据表\\'
    # print(month_data_base_path)
    all_month_tables = take_tables(month_data_base_path)
    table_to_sheet(all_month_tables, sheet_finaly_workbook_month)
    finaly_workbook_month.save(filename=sys.path[0] + '\\表格\\月平衡计分卡.xlsx')
    print('月数据全部填写完成！')


if __name__ == "__main__":
    week_data()
    # month_data()
