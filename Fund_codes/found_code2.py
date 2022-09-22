import os
import re
import json
import time
import warnings
import requests
import datetime
import openpyxl   # versiong==2.4.9，使用3.0.9、3.0.10读取excel表格会报错
from openpyxl import load_workbook

warnings.filterwarnings('ignore')


# 浏览器头
headers = {'content-type': 'application/json; charset=utf-8',
           'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.0.0 Safari/537.36'}


# 获取预估数据 或 历史数据
def get_data(url):
    r = requests.get(url, headers=headers)
    # 返回信息
    content = r.text
    # 正则表达式
    pattern = r'^jsonpgz\((.*)\)'
    # 查找结果
    search = re.findall(pattern, content)
    data = json.loads(search[0])
    
    return  data


# 写入历史数据和预估数据
def write_history_data(estimate_data, sheet):
    latestDate = estimate_data['jzrq']
    nowDate = estimate_data['gztime'][:10]
    
    for date_ in [latestDate, nowDate]:
        if date_ not in all_date:
            sheet.insert_rows(idx=3, amount=1)
            row = 3
        else:
            # 获取A列数据
            is_A = [i.value for i in sheet['A']]
            # 获取A列数据为1（买入点）的索引
            is_A_index = [i for i,v in enumerate(is_A) if v==date_]
            row = is_A_index[0] + 1
        
        # 日期
        sheet.cell(row=row, column=1).value = date_
        # 净值
        if date_==latestDate:
            sheet.cell(row=row, column=2).value = float(estimate_data['dwjz'])
        else:
            sheet.cell(row=row, column=2).value = float(estimate_data['gsz'])
        # 涨跌幅
        sheet.cell(row=row, column=3).value = round(((sheet.cell(row=row, column=2).value - sheet.cell(row=row+1, column=2).value) / sheet.cell(row=row+1, column=2).value) * 100, 2)
    
    
# 计算累计涨跌幅
def whetherNotToBuy(sheet):
    # 获取E列数据
    is_E = [i.value for i in sheet['E']]
    # 获取D列数据为1（买入点）的索引
    is_E_index = [i for i,v in enumerate(is_E) if v==1]
    row = is_E_index[0]   # 上次购买所在的行数的前一行
    # 累计涨跌幅度
    for i in range(row-2): # 减去两行表头
        sheet[f'D{row-i}'] = sum([j.value for j in sheet['C'][row-i-1:row]])  # row-i-1是由于sheet中切片是前闭后开的，所以要减去一行
        if sheet[f'D{row-i}'].value >= -4:
            sheet[f'E{row-i}'] = 0
        else:
            sheet[f'E{row-i}'] = 1
        # print(sheet[f'E{row}'].value)
        
        
codes = ['161725', '005827', '003095']
# codes = ['005827']

# 获取表格对象
workbook = load_workbook('/Users/ayd/Desktop/Github_Repositorys/MyOwn/My_tools/Fund_codes/基金数据买卖点.xlsx')
sheetnames = workbook.sheetnames

for code in codes:
    # 生成链接
    url = "http://fundgz.1234567.com.cn/js/%s.js"%code
    
    # 获取数据
    estimate_data = get_data(url)
    
    ##### 数据入表 #####
    # 创建sheet
    if estimate_data['name'] not in sheetnames:
        workbook.create_sheet(estimate_data['name'])
    sheet = workbook[estimate_data['name']]

    # 表头写入
    sheet['A1'] = '基金代码：' + code
    sheet['A2'] = '交易日期'
    sheet['B2'] = '单位净值'
    sheet['C2'] = '净值涨跌'
    sheet['D2'] = '自上次买入累计涨跌'
    sheet['E2'] = '是否买入'
    sheet['F2'] = '买入金额'    

    # 表中已有的数据的所有日期
    all_date = [sheet['A'][a].value for a in range(len(sheet['A']))][2:]
    
    # 写入历史数据
    write_history_data(estimate_data, sheet)
    # 计算距离上次买入的累计涨跌
    whetherNotToBuy(sheet)

workbook.save('/Users/ayd/Desktop/Github_Repositorys/MyOwn/My_tools/Fund_codes/基金数据买卖点.xlsx')

print('完成')